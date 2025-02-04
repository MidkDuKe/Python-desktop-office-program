import os
import sys
from tkinter.simpledialog import askstring

import openpyxl
import pandas as pd
import ttkbootstrap as ttk  # 替换导入tkinter
from ttkbootstrap.constants import *  # 导入ttkbootstrap常量
from tkinter import messagebox, filedialog
from openpyxl import load_workbook
from PIL import Image, ImageTk


def save_and_exit(entries, tree):
    global initial_data
    need_path=file_entry.get()
    initial_data = {key: entry.get() for key, entry in entries.items()}
    df = pd.DataFrame([initial_data])
    df.loc[0, '路径'] = need_path
    selected_data = pair_df[['学号', '工时']]
    with pd.ExcelWriter(need_path) as writer:
        selected_data.to_excel(writer, index=False, sheet_name='Sheet1')
    with pd.ExcelWriter("./data/Setting.xlsx") as writer:
        df.to_excel(writer, index=False)
    main()
    messagebox.showinfo("信息", "输出成功！")



def open_new_window():
    new_window = ttk.Toplevel(root)
    new_window.title("帮助")
    new_window.geometry("870x320")
    new_window.resizable(False, False)
    new_window.iconbitmap("data/images/check.ico")

    frame = ttk.Frame(new_window)
    frame.pack(padx=10, pady=10)

    image_path = "data/images/weixinpay.png"
    try:
        image = Image.open(image_path)
        image = image.resize((250, 250), Image.LANCZOS)
        photo = ImageTk.PhotoImage(image)
        img_label = ttk.Label(frame, image=photo)
        img_label.image = photo
        img_label.grid(row=0, column=0, padx=10, pady=10)
    except FileNotFoundError:
        messagebox.showerror("错误", f"未找到图片: {image_path}")

    text_content = """
    欢迎使用本软件(●’◡’●)可以扫描左侧二维码对作者进行打赏，如果有问题可以发送邮件至midk_duke@qq.com
    软件功能：
        1.选择文件：选取的文件需要和根目录的输入.xlsx格式一致，若报错无法打开，请在data/Setting.xlsx文件中修改路径。
        2.打开文件：可以通过文件修改，但不要修改数据格式，也可以直接双击数据进行修改。
        3.添加：输入学号和工时会自动补全姓名。
        4.删除：选中后点击即可删除。
        5.运行：配置和数据将会保存，并输出结果，点击结果按钮查看。
    提示：
        1.暂不支持加班功能，可以在输出结果中手动修改。
        2.在第一次运行时，请先点击信息库按钮将所有可能包含的人员信息导入。
    """
    text_label = ttk.Label(frame, text=text_content, justify=LEFT, wraplength=565)
    text_label.grid(row=0, column=1, padx=10, pady=5)


# 读取 Excel 文件
def readdict():
    global info_dict
    info_df = pd.read_excel('./data/信息.xlsx', sheet_name='Sheet1', names=['姓名', '学号'])
    info_dict = dict(zip(info_df['学号'], info_df['姓名']))
readdict()
check_df = pd.read_excel('./data/Setting.xlsx')
Department=check_df.iloc[0,0]
year=check_df.iloc[0,1]
month=check_df.iloc[0,2]
if month<10:
    monthStr='0'+str(month)
else:
    monthStr=str(month)
end=check_df.iloc[0,3]
when=check_df.iloc[0,4]
CheckMan=check_df.iloc[0,5]
Manager=check_df.iloc[0,6]
money=check_df.iloc[0,7]
account=check_df.iloc[0,8]
need_path=check_df.iloc[0,9]
# 初始化输入数据
initial_data = {
    "部门/岗位": Department,
    "年份": year,
    "汇总月份": month,
    "该月天数": end,
    "制表日期": when,
    "制表人": CheckMan,
    "负责人": Manager,
    "时薪": money,
    "账户": account,
    "路径": need_path
}
# 创建一个学号到姓名的字典

pair_df=pd.DataFrame()
failed_pairs = []
def readnewpath():
    global pair_df
    global failed_pairs
    global need_path
    failed_pairs.clear()
    pair_df.drop(pair_df.index, inplace=True)
    pair_df= pd.read_excel(need_path, names=['学号', '工时'])
    # 在配对表中根据学号添加姓名数据，并检查是否配对失败
    pair_df['姓名'] = pair_df['学号'].apply(lambda x: info_dict.get(x, ''))
    # 检查哪些学号没有匹配到姓名
    failed_pairs = pair_df[pair_df['姓名'] == '']['学号'].tolist()
readnewpath()
# 创建主窗口
root = ttk.Window(themename="litera")  # 设置ttkbootstrap主题
root.title("Check And Insert V2.3.1 by DuKe")
root.iconbitmap('data/images/check.ico')
root.configure(padx=10, pady=10)

frame_left = ttk.Labelframe(root,text="配置",bootstyle='primary')
frame_left.grid(row=0, column=0, rowspan=10, padx=10, pady=10, sticky="nsew")

entries = {}
for i, (key, value) in enumerate(initial_data.items()):
    if i >= 9:
        break
    ttk.Label(frame_left, text=key).grid(row=i, column=0, padx=10, pady=5, sticky='e')
    entry = ttk.Entry(frame_left, width=25)
    entry.grid(row=i, column=1, padx=10, pady=5)
    entry.insert(0, str(value))
    entries[key] = entry

control_frame = ttk.Frame(root)
control_frame.grid(row=0, column=1,rowspan=10, padx=10, pady=10, sticky="ew")
def choose_file():
    global need_path
    file_path = filedialog.askopenfilename(title="文件路径", filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
    if file_path:
        need_path = file_path
        file_entry.delete(0, 'end')
        file_entry.insert(0, need_path)
    readnewpath()
    insert_toTree()

choose_frame = ttk.Frame(control_frame)
choose_frame.grid(row=0, column=0, columnspan=8, padx=10, pady=10, sticky="ew")

file_entry = ttk.Entry(choose_frame, width=52)
file_entry.insert(0, need_path)
file_entry.grid(row=0, column=1, padx=10, pady=5)
choose_button = ttk.Button(choose_frame, text="选择文件", command=choose_file, width=9)
choose_button.grid(row=0, column=0, padx=10, pady=5)


columns = ("姓名", "学号", "工时")
tree_frame = ttk.Frame(control_frame)
tree_frame.grid(row=2, column=0, columnspan=4, padx=20, pady=5, rowspan=len(initial_data), sticky="nsew")
tree = ttk.Treeview(tree_frame, columns=columns, show='headings')
tree.heading("学号", text="学号")
tree.heading("姓名", text="姓名")
tree.heading("工时", text="工时")
tree.grid(row=0, column=0, sticky="nsew")
scrollbar = ttk.Scrollbar(tree_frame, orient=VERTICAL, command=tree.yview,bootstyle='primary')
tree.configure(yscroll=scrollbar.set)
scrollbar.grid(row=0, column=1, sticky='ns')
def insert_toTree():
    tree.delete(*tree.get_children())
    if failed_pairs:
        messagebox.showinfo("警告", "以下学号配对失败，请检查输入是否正确：\n" + "\n".join(failed_pairs))
        sys.exit()

    for index, row in pair_df.iterrows():
        tree.insert("", "end", values=(row["姓名"], row["学号"], row["工时"]))
insert_toTree()

def on_double_click(event):
    item = tree.selection()[0]
    column_id = tree.identify_column(event.x)
    if column_id:
        column_name = tree.heading(column_id)["text"]
        value = tree.item(item, 'values')[columns.index(column_name)]
        new_value = askstring("编辑", f"请输入新的{column_name}: ", initialvalue=value)
        if new_value is not None:
            tree.set(item, column_name, new_value)
            index = int(tree.index(item))
            if column_name in ["学号"]:
                new_value = int(new_value)
                pair_df.loc[index, column_name] = new_value
            if column_name in ["工时"]:
                new_value = float(new_value)
                pair_df.loc[index, column_name] = new_value

tree.bind('<Double-1>', on_double_click)

se_frame = ttk.Frame(control_frame)
se_frame.grid(row=1, column=0, columnspan=8, padx=10, pady=10, sticky="ew")

save_button = ttk.Button(se_frame, text="运行", command=lambda: save_and_exit(entries, tree), width=9,bootstyle="success")
save_button.grid(row=0, column=0, padx=10, pady=5)
open_button = ttk.Button(se_frame, text="结果", command=lambda: open_file(), width=9)
open_button.grid(row=0, column=1, padx=10, pady=5)
edit_button=ttk.Button(se_frame, text="打开文件", command=lambda: open_excel(),width=9)
info_button=ttk.Button(se_frame, text="信息库", command=lambda: open_info(),width=9)
edit_button.grid(row=0, column=2,  padx=10, pady=5)
info_button.grid(row=0, column=3, padx=10, pady=5)
help_button = ttk.Button(se_frame, text="帮助", command=open_new_window, width=9)
help_button.grid(row=0, column=4, padx=10, pady=5)
def open_info():
    os.startfile(r'data\信息.xlsx')
    messagebox.showinfo("提示", "请在修改并保存后重新启动软件，只需要修改一次！")
def open_excel():
    os.startfile(need_path)
    messagebox.showinfo("提示", "请在修改并保存后重新选择文件！")
def open_file():
    os.startfile('out')
def add_data(tree, pair_df, student_id_entry, hours_entry):

    # 获取输入的学号和工时
    student_id = int(student_id_entry.get())
    hours = hours_entry.get()
    if not hours.isdigit():
        messagebox.showerror("警告", "工时必须为数字！")
        return
    # 获取学号对应的姓名，如果学号不存在则返回空字符串
    name = info_dict.get(student_id, '')
    hours=float(hours)
    # 检查是否成功匹配到姓名
    if not name:
        messagebox.showwarning("警告", f"学号 {student_id} 在信息表中不存在，请检查输入！")
        return
    # 将数据添加到 Treeview 和 pair_df 中
    tree.insert("", "end", values=(name, student_id, hours ))
    pair_df.loc[len(pair_df)] = [student_id, hours, name]
    print(f"成功添加学号为 {student_id} 的姓名为 {name} 的数据")

# 创建添加按钮
add_frame = ttk.Frame(control_frame)
add_frame.grid(row=15, column=0, padx=10, pady=5)

# 创建输入框和标签
add_student_id_label = ttk.Label(add_frame, text="学号")
add_student_id_label.grid(row=0, column=0, padx=5, pady=5)
add_student_id = ttk.Entry(add_frame,width=13)
add_student_id.grid(row=0, column=1, padx=5, pady=5)
add_hours_label = ttk.Label(add_frame, text="工时")
add_hours_label.grid(row=0, column=2, padx=5, pady=5)
add_hours = ttk.Entry(add_frame,width=13)
add_hours.grid(row=0, column=3, padx=10, pady=5)
add_button = ttk.Button(add_frame, text="添加", command=lambda: add_data(tree, pair_df, add_student_id, add_hours),width=9)
add_button.grid(row=0, column=4, padx=10, pady=5)
delete_button = ttk.Button(add_frame, text="删除", command=lambda: delete_row(tree, pair_df),width=9,bootstyle="danger")
delete_button.grid(row=0, column=5, padx=10, pady=5)
def delete_row(tree, pair_df):
    selected_items = tree.selection()
    if not selected_items:
        messagebox.showwarning("警告", "请选择要删除的行！")
        return
    for item in selected_items:
        # 获取要删除的行的学号
        student_id = tree.item(item)["values"][1]
        # 从 Treeview 中删除选中的行
        tree.delete(item)
        # 从 pair_df 中删除对应的行
        pair_df.drop(pair_df[pair_df["学号"] == student_id].index, inplace=True)
        print(f"成功删除学号为 {student_id} 的数据")
# 创建删除按钮，并绑定删除函数
root.resizable(False, False)

def main():
    global pair_df
    global initial_data
    Department = initial_data['部门/岗位']
    year = initial_data['年份']
    month = int(initial_data['汇总月份'])
    if month < 10:
        monthStr = '0' + str(month)
    else:
        monthStr = str(month)
    end = int(initial_data['该月天数'])
    when = initial_data['制表日期']
    CheckMan = initial_data['制表人']
    Manager = initial_data['负责人']
    money = int(initial_data['时薪'])
    account = initial_data['账户']
    # 读取格式.xlsx和配对.xlsx
    format_df = pd.read_excel('data/xx助理xx月xx日进卡工资（新）.xlsx')
    format_df2 = pd.read_excel('data/xx助理xx月xx日进卡工资（旧）.xlsx')
    pair_df = pair_df[['学号', '姓名', '工时']]
    pair_edit=pair_df.copy()
    pair_edit['学号']=pair_df['学号'].astype(str)
    pair_df2 = pair_edit.copy()
    # 计算工资并四舍五入到指定位数（比如两位小数）
    rounded_wages = (pair_df2['工时'] * money).round(1)
    # 将四舍五入后的工资赋值给'工时'列（注意这可能会覆盖原始数据）
    pair_df2['工时'] = rounded_wages

    # 加载格式.xlsx文件
    workbook = load_workbook('data/xx助理xx月xx日进卡工资（新）.xlsx')
    sheet = workbook.active
    workbook2 = load_workbook('data/xx助理xx月xx日进卡工资（旧）.xlsx')
    sheet2 = workbook2.active

    # 从第6行开始写入配对表数据，第2列到第4列
    start_row = 6
    start_row2 = 4
    # 写入配对表数据
    for index, row in pair_edit.iterrows():
        for col_index, value in enumerate(row, start=2):
            sheet.cell(row=start_row + index, column=col_index, value=value)

    for index, row in pair_df2.iterrows():
        for col_index, value in enumerate(row, start=2):
            sheet2.cell(row=start_row2 + index, column=col_index, value=value)

    # 获取编号列的所有值
    id_values = format_df.iloc[:, 0].values

    # 获取配对表中的学号
    written_ids = pair_df['学号'].values

    # 删除编号为1-80范围内学号为空的数据，并保留表尾的汇总行
    rows_to_delete = []
    for row in range(start_row, start_row + 80):
        cell_value = sheet.cell(row=row, column=2).value  # 第二列为学号列
        if not cell_value:  # 如果学号为空
            rows_to_delete.append(row)
    rows_to_delete2 = []
    for row in range(start_row2, start_row2 + 80):
        cell_value = sheet2.cell(row=row, column=2).value  # 第二列为学号列
        if not cell_value:  # 如果学号为空
            rows_to_delete2.append(row)
    # 删除未写入数据的行，保留表尾汇总行
    for row in reversed(rows_to_delete):
        sheet.delete_rows(row, 1)
    for row in reversed(rows_to_delete2):
        sheet2.delete_rows(row, 1)
    # 查找最大编号行，确定end_row
    max_row = 0
    for row in range(start_row, start_row + 80):
        cell_value = sheet.cell(row=row, column=1).value  # 第一列为编号列
        if cell_value is None:
            break
        max_row = row

    max_row2 = 0
    for row in range(start_row2, start_row2 + 80):
        cell_value = sheet2.cell(row=row, column=1).value  # 第一列为编号列
        if cell_value is None:
            break
        max_row2 = row

    for row in range(start_row, max_row):
        # 构造F列和G列的单元格对象
        cell_f = sheet.cell(row=row, column=6, value=None)  # F列的列号是6
        cell_g = sheet.cell(row=row, column=7)  # G列的列号是7
        # 设置G列单元格的公式为F列单元格的值乘以21
        cell_g.value = f'={cell_f.coordinate}*{money}'

    def convert_to_rmb_upper(amount):
        # 定义中文数字和单位
        CHINESE_NUMBERS = ['零', '壹', '贰', '叁', '肆', '伍', '陆', '柒', '捌', '玖']
        CHINESE_BIG_UNITS = ['', '拾', '佰', '仟', '万', '拾', '佰', '仟', '亿']
        DECIMAL_UNITS = ['角', '分']

        # 确保输入为合法的数字
        if not isinstance(amount, (int, float)):
            raise ValueError('输入必须为整数或浮点数')

            # 分离整数部分和小数部分
        integer_part = int(amount)
        decimal_part = round((amount - integer_part) * 100)  # 将小数转换为整数部分（0-99）

        # 转换整数部分
        chinese_integer = ''
        is_zero = False  # 标记连续零的状态
        for i in range(len(str(integer_part)) - 1, -1, -1):
            digit = int(str(integer_part)[i])
            unit = CHINESE_BIG_UNITS[len(str(integer_part)) - 1 - i]
            if digit == 0:
                if not is_zero and chinese_integer:
                    chinese_integer = CHINESE_NUMBERS[0] + chinese_integer
                    is_zero = True
            else:
                chinese_integer = CHINESE_NUMBERS[digit] + unit + chinese_integer
                is_zero = False

                # 去除末尾不必要的零
        chinese_integer = chinese_integer.rstrip('零')

        # 转换小数部分
        chinese_decimal = ''
        if decimal_part > 0:
            jiao = decimal_part // 10
            fen = decimal_part % 10
            if jiao:
                chinese_decimal += CHINESE_NUMBERS[jiao] + DECIMAL_UNITS[0]
            if fen:
                chinese_decimal += CHINESE_NUMBERS[fen] + DECIMAL_UNITS[1]
        elif decimal_part == 0:
            chinese_decimal += '整'

            # 拼接结果
        if chinese_integer or chinese_decimal:
            result = chinese_integer + '元' + chinese_decimal
        else:
            result = '零元整'

        return result

    # 查找汇总行“小计”并填入SUM公式
    def insert_summary_formula(sheet, summary_row_label='小计', start_row=6, end_row=None, start_col='D', end_col='L'):
        # 查找汇总行“小计”
        summary_row = None
        for row in sheet.iter_rows(min_row=start_row, max_row=sheet.max_row, min_col=1, max_col=1):
            if row[0].value == summary_row_label:
                summary_row = row[0].row
                break

        if summary_row is None:
            raise ValueError(f"未找到汇总行'{summary_row_label}'")

        # 填入SUM公式
        for col in range(ord(start_col), ord(end_col) + 1):
            col_letter = chr(col)
            formula = f"=SUM({col_letter}{start_row}:{col_letter}{end_row})"
            sheet[f"{col_letter}{summary_row}"].value = formula

        # 合并汇总行的A到C列
        sheet.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=3)

    def insert_summary_formula2(sheet, summary_row_label='小计', start_row=6, end_row=None, start_col='D', end_col='L'):
        # 查找汇总行“小计”
        summary_row = None
        for row in sheet.iter_rows(min_row=start_row, max_row=sheet.max_row, min_col=1, max_col=1):
            if row[0].value == summary_row_label:
                summary_row = row[0].row
                break

        if summary_row is None:
            raise ValueError(f"未找到汇总行'{summary_row_label}'")

        # 填入SUM公式
        for col in range(ord(start_col), ord(end_col) + 1):
            col_letter = chr(col)
            formula = f"=SUM({col_letter}{start_row}:{col_letter}{end_row})"
            sheet[f"{col_letter}{summary_row}"].value = formula
            sheet[f"{col_letter}{summary_row}"].number_format = '¥#,##0.00;¥-#,##0.00'
            pair_df2['工时']=pair_df2['工时'].astype(float)
            sum_value = pair_df2['工时'].sum()
            # for row in range(4, summary_row):  # 假设第一行是1（在openpyxl中通常是这样）
            #     cell_value = sheet.cell(row=row, column=4).value
            #     if cell_value is not None:  # 确保单元格有值（不是None）
            #         sum_value += cell_value
            sheet[f"{chr(col - 2)}{summary_row}"].value = convert_to_rmb_upper(sum_value)
            sheet[f"{chr(col - 2)}{summary_row + 2}"].value = CheckMan
            if sheet[f"{chr(col + 1)}{summary_row - 4}"].value=='部门':
                sheet[f"{chr(col + 1)}{summary_row - 3}"].value = Department+'(' + str(
                    month) + '.01-' + str(month) + '.' + str(end) + ')'
                sheet[f"{chr(col + 1)}{summary_row - 2}"].value = formula
                sheet[f"{chr(col + 1)}{summary_row - 2}"].number_format = '¥#,##0.00;¥-#,##0.00'
                sheet[f"{chr(col + 1)}{summary_row - 1}"].value = "制表人：" + CheckMan
            elif sheet[f"{chr(col + 1)}{summary_row - 3}"].value=='部门':
                sheet[f"{chr(col + 1)}{summary_row - 2}"].value = Department+'(' + str(
                    month) + '.01-' + str(month) + '.' + str(end) + ')'+"制表人：" + CheckMan
                sheet[f"{chr(col + 1)}{summary_row - 1}"].value = formula
                sheet[f"{chr(col + 1)}{summary_row - 1}"].number_format = '¥#,##0.00;¥-#,##0.00'
            elif sheet[f"{chr(col + 1)}{summary_row - 2}"].value=='部门':
                sheet[f"{chr(col + 1)}{summary_row - 1}"].value = Department+'(' + str(
                    month) + '.01-' + str(month) + '.' + str(end) + ')'+'¥'+str(sum_value)+"制表人：" + CheckMan
            else:
                sheet[f"{chr(col + 1)}{summary_row - 2}"].value = formula
                sheet[f"{chr(col + 1)}{summary_row - 2}"].number_format = '¥#,##0.00;¥-#,##0.00'
                sheet[f"{chr(col + 1)}{summary_row - 1}"].value = "制表人：" + CheckMan
                if (summary_row - 4) % 2 == 0:
                    sheet[f"{chr(col + 1)}{int(summary_row - 2 - (summary_row - 4) / 2)}"].value = Department
                    sheet[f"{chr(col + 1)}{int(summary_row -1- (summary_row - 4) / 2)}"].value = '(' + str(
                        month) + '.01-' + str(month) + '.' + str(end) + ')'
                else:
                    sheet[f"{chr(col + 1)}{int(summary_row - 2 - (summary_row - 5) / 2)}"].value = Department
                    sheet[f"{chr(col + 1)}{int(summary_row - 1-(summary_row - 5) / 2)}"].value = '(' + str(
                        month) + '.01-' + str(month) + '.' + str(end) + ')'

        # 合并汇总行的A到C列
        sheet.merge_cells(start_row=summary_row, start_column=2, end_row=summary_row, end_column=3)

    # 插入汇总行公式
    insert_summary_formula(sheet, summary_row_label='小计', start_row=6, end_row=max_row - 1, start_col='D',
                           end_col='L')
    insert_summary_formula2(sheet2, summary_row_label='总计', start_row=4, end_row=max_row2 - 3, start_col='D',
                            end_col='D')

    def update_cell_in_xlsx(sheet_index, cell_position, new_content):
        sheet[cell_position] = new_content
        print(f"指定数据'{new_content}'已填入！")

    def update_cell_in_xlsx2(sheet_index, cell_position, new_content):
        sheet2[cell_position] = new_content
        print(f"指定数据'{new_content}'已填入！")

    update_cell_in_xlsx(0, 'A1', '勤工助学指导中心学生' + str(month) + '月份工资核算表')
    update_cell_in_xlsx(0, 'A2', '工资日期范围:' + str(year) + '/' + monthStr + '/01-' + str(year) + '/' + str(
        month) + '/' + str(end))
    update_cell_in_xlsx(0, 'F2', '制表日期：' + str(year) + '年' + when)
    update_cell_in_xlsx(0, 'J2', '部门:' + Department)
    update_cell_in_xlsx(0, 'A3',
                        '账号：' + account + '     制表人：' + CheckMan + '          部门负责人：' + Manager + '      财务负责人：###          财务老师：####    	')
    workbook.save('out/新表/' + Department + str(month + 1) + '月15日进卡工资（新）.xlsx')

    update_cell_in_xlsx2(0, 'A1', '勤工助学指导中心' + str(month + 1) + '月15日进卡工资（奉贤' + account + '）')
    update_cell_in_xlsx2(0, 'A2', '制表时间：' + str(year) + '年' + when)
    workbook2.save('out/旧表/' + Department + str(month + 1) + '月15日进卡工资（旧）.xlsx')
    print("录入完成！")
    pair_df = pair_df[['学号','工时','姓名']]

# 运行主循环
root.mainloop()
